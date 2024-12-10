import os
import openpyxl
from datetime import datetime
import openpyxl.workbook
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import NotFound
from rest_framework import status
from rest_framework import generics
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Optional, cast

from server import settings
from api.models import Project, TestPlan, TestCase, TestCoverage, DefectReport, TestReport
from api.serializers import (
    ProjectSerializer,
    TestPlanSerializer,
    TestCaseSerializer,
    TestCoverageSerializer,
    DefectReportSerializer,
    TestReportSerializer,
)

# tools for excel sheet

border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(
    start_color="22f071", end_color="22f071", fill_type="solid")
header_font = Font(bold=True)
header_alignment = Alignment(horizontal="center", vertical="center")
cell_alignment_wrap = Alignment(
    wrap_text=True, vertical="top", horizontal="left")
cell_alignment_center = Alignment(horizontal="center", vertical="center")


# handle project list
# Created date    :   26/11/2024


class ProjectListView(generics.ListCreateAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


# handle project detail [retrieve-update-destory]
# Created date    :   26/11/2024


class ProjectDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]


# handle testplan [list/create]
# Created date    :   26/11/2024


class TestPlanListView(generics.ListCreateAPIView):
    queryset = TestPlan.objects.all()
    serializer_class = TestPlanSerializer
    permission_classes = [IsAuthenticated]

# handle Testplan detail [retrive-update-destory/export]
# Created date    :   26/11/2024
# Updated date    :   04/12/2024


class TestPlanDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TestPlan.objects.all()
    serializer_class = TestPlanSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        if 'export' in request.query_params:
            return self.export_file(request, *args, **kwargs)

        return super().get(request, *args, **kwargs)

    def export_file(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        data: Dict[str, Optional[str]] = cast(
            Dict[str, Optional[str]], response.data)

        project_id: Optional[str] = data.get('project')
        project_name = Project.objects.filter(id=project_id).values_list(
            'name', flat=True).first() if project_id else None

        data['project'] = project_name or f"Unknown Project (ID: {project_id})"

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None, "Failed"

        sheet.title = 'Test Plan'

        headers = ["Sections", "Description"]

        sheet.append(headers)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        sections = [
            ("Project", data.get('project')),
            ("Objective", data.get('objective')),
            ("Scope In", data.get('scope_in')),
            ("Scope Out", data.get('scope_out')),
            ("Test Levels", data.get('test_levels')),
            ("Types of Testing", data.get('types_of_testing')),
            ("Environment Details", data.get('environment_details')),
            ("Test Data", data.get('test_data')),
            ("Test Manager", data.get('test_manager')),
            ("Test Leads", data.get('test_leads')),
            ("Testers", data.get('testers')),
            ("Developers", data.get('developers')),
            ("Business Analysts", data.get('business_analysts')),
            ("Milestones", data.get('milestones')),
            ("Deadlines", data.get('deadlines')),
            ("Dependencies", data.get('dependencies')),
            ("Deliverables", data.get('deliverables')),
            ("Entry Criteria", data.get('entry_criteria')),
            ("Exit Criteria", data.get('exit_criteria')),
            ("Risks", data.get('risks')),
            ("Mitigation Strategies", data.get(
                'mitigation_strategies')),
            ("Defect Management", data.get('defect_management')),
            ("Communication Plan", data.get('communication_plan')),
            ("Approval Process", data.get('approval_process')),
            ("Sign-off Authorities", data.get('sign_off_authorities')),
        ]

        for section, description in sections:
            sheet.append([section, description])

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

            for cell in col:
                cell.alignment = Alignment(wrap_text=True)
                if cell.value:
                    cell.border = border_style

        current_time = datetime.now().strftime("%Y%m%d-%H%M%S")
        media_root = settings.MEDIA_ROOT
        download_path = os.path.join(media_root, 'download')
        os.makedirs(download_path, exist_ok=True)  # Ensure the folder exists

        file_name = f'testplan-{current_time}.xlsx'
        file_path = os.path.join(download_path, file_name)

        # Save the workbook to the file path
        workbook.save(file_path)

        # Respond with the file location or success message
        return Response({
            'message': 'File has been successfully saved.',
            'file_path': file_path,
        })


# handle Testcase [list/create/export]
# Created date  : 26/11/2024
# updated date  : 05/12/2024


class TestCaseListView(generics.ListCreateAPIView):
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    permission_classes = [IsAuthenticated]\


    def list(self, request, *args, **kwargs):
        if 'export' in request.query_params:
            return self.export_file(request, *args, **kwargs)

        return super().list(request, *args, **kwargs)

    def export_file(self, request, pk=None, *args, **kwargs):
        project_id = pk
        project_name = Project.objects.filter(id=project_id).values_list(
            'name', flat=True).first() if project_id else None

        if not project_id:
            return Response({"detail": "Project ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        test_cases = TestCase.objects.filter(project_id=pk)
        if not test_cases.exists():
            return Response({"detail": "No test cases found for this project"}, status=status.HTTP_404_NOT_FOUND)

        mapped_ids = {test_case.testcaseID: f"TC_{idx:04d}" for idx,
                      test_case in enumerate(test_cases, start=1)}

        wb = openpyxl.Workbook()
        sheet = wb.active
        assert sheet is not None, "Failed"

        sheet.title = "Test Cases"

        headers = ["S.no", "Test Case ID", "Description", "Pre-conditions",
                   "Test Steps", "Expected Results", "Actual Result", "Status"]
        sheet.append(headers)

        wrap_columns = ["Description", "Pre-conditions",
                        "Test Steps", "Expected Results", "Actual Result"]

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style

        column_widths = [6, 15, 50, 50, 50, 50, 50, 15]
        for i, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = width

        wrap_columns = {"C", "D", "E", "F", "G"}

        for idx, test_case in enumerate(test_cases, start=1):
            row_data = [
                idx,  # Serial number
                # test_case.testcaseID,
                mapped_ids[test_case.testcaseID],
                test_case.description,
                test_case.preconditions,
                test_case.test_steps,
                test_case.expected_result,
                test_case.actual_result,
                test_case.status
            ]
            sheet.append(row_data)

            for col_idx, cell in enumerate(sheet[idx + 1], start=1):
                cell.border = border_style
                if cell.column_letter in wrap_columns:
                    cell.alignment = cell_alignment_wrap
                else:
                    cell.alignment = cell_alignment_center if col_idx <= 2 else Alignment(
                        horizontal="left")

        current_time = datetime.now().strftime("%Y%m%d-%H%M%S")
        media_root = settings.MEDIA_ROOT
        download_path = os.path.join(media_root, 'download')
        os.makedirs(download_path, exist_ok=True)

        file_name = f'{project_name}_testcase-{current_time}.xlsx'
        file_path = os.path.join(download_path, file_name)

        wb.save(file_path)

        return Response({
            'message': 'File has been successfully saved.',
            'file_path': file_path,
        })


# handle Testcase detail [retrieve-update-destory]
# Created date    :   26/11/2024


class TestCaseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    permission_classes = [IsAuthenticated]


# handle Testcoverage [list/create]
# Created date    :   26/11/2024

class TestCoverageListView(generics.ListCreateAPIView):
    queryset = TestCoverage.objects.all()
    serializer_class = TestCoverageSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except TestCase.DoesNotExist:
            raise NotFound(
                detail="TestCase does not exist, please verify your input data.")


# handle Testcoverage detail [retrieve-update-destory]
# Created date    :   09/12/2024


class TestCoverageDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TestCoverage.objects.all()
    serializer_class = TestCoverageSerializer
    permission_classes = [IsAuthenticated]


# handle defectreport [list/create/export]
# Created date    :   04/12/2024
# Updated date    :   10/12/2024


class DefectReportListView(generics.ListCreateAPIView):
    queryset = DefectReport.objects.all()
    serializer_class = DefectReportSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        if 'export' in request.query_params:
            return self.export_file(request, *args, **kwargs)

        return super().list(request, *args, **kwargs)

    def export_file(self, request, pk=None, *args, **kwargs):
        project_id = pk
        project_name = Project.objects.filter(id=project_id).values_list(
            'name', flat=True).first() if project_id else None

        if not project_id:
            return Response({'details': 'Project ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        defect_reports = DefectReport.objects.filter(project_id=pk)
        if not defect_reports.exists():
            return Response({'details': 'No defect report found for this project'}, status=status.HTTP_404_NOT_FOUND)

        mapped_ids = {defect_report.defect_id: f"DEFECT_{idx:04d}" for idx,
                      defect_report in enumerate(defect_reports, start=1)}

        wb = openpyxl.Workbook()
        sheet = wb.active
        assert sheet is not None, "Failed"

        sheet.title = "Defect Report"

        headers = ["S.no", "Defect ID", "Summary",
                   "Steps to Reproduce", "Severity", "Status", "Evidence", "Defect Detected Date", "Defect Fixed Date", "Re-open Defect Date"]
        sheet.append(headers)

        wrap_columns = ["Summary", "Steps to Reproduce", "Evidence"]

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style

        column_widths = [6, 15, 50, 50, 15, 15, 40, 25, 25, 25]
        for i, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = width

        wrap_columns = {"C", "D", "G"}

        for idx, defect_report in enumerate(defect_reports, start=1):
            evidence_url = defect_report.evidence.url if defect_report.evidence else None

            defect_detected_date = defect_report.defect_detected_date.replace(
                tzinfo=None) if defect_report.defect_detected_date else ''
            defect_fixed_date = defect_report.defect_fixed_date.replace(
                tzinfo=None) if defect_report.defect_fixed_date else ''
            reopen_defect_date = defect_report.reopen_defect_date.replace(
                tzinfo=None) if defect_report.reopen_defect_date else ''

            row_data = [
                idx,  # Serial number
                mapped_ids[defect_report.defect_id],
                defect_report.summary,
                defect_report.steps_to_reproduce,
                defect_report.severity,
                defect_report.status,
                evidence_url,
                defect_detected_date,
                defect_fixed_date,
                reopen_defect_date
            ]
            sheet.append(row_data)

            for col_idx, cell in enumerate(sheet[idx + 1], start=1):
                cell.border = border_style
                col_letter = get_column_letter(col_idx)

                if col_letter in wrap_columns:
                    cell.alignment = cell_alignment_wrap
                elif col_letter in ['E', 'F', 'G', 'H', 'I', 'J']:
                    cell.alignment = cell_alignment_center
                else:
                    cell.alignment = cell_alignment_center if col_idx <= 2 else Alignment(
                        horizontal="left")

        current_time = datetime.now().strftime("%Y%m%d-%H%M%S")
        media_root = settings.MEDIA_ROOT
        download_path = os.path.join(media_root, 'download')
        os.makedirs(download_path, exist_ok=True)

        file_name = f'{project_name}_DefectReport-{current_time}.xlsx'
        file_path = os.path.join(download_path, file_name)

        wb.save(file_path)

        return Response({
            'message': 'File has been successfully saved.',
            'file_path': file_path,
        })


class DefectReportDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DefectReport.objects.all()
    serializer_class = DefectReportSerializer
    permission_classes = [IsAuthenticated]


# handle Testreport [list/create]
# Created date    :   04/12/2024
# Updated date    :   10/12/2024


class TestreportListView(generics.ListCreateAPIView):
    queryset = TestReport.objects.all()
    serializer_class = TestReportSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        if 'export' in request.query_params:
            return self.export_file(request, *args, **kwargs)

        return super().list(request, *args, **kwargs)

    def export_file(self, request, pk=None, *args, **kwargs):
        project_id = pk
        project_name = Project.objects.filter(id=project_id).values_list(
            'name', flat=True).first() if project_id else None

        if not project_id:
            return Response({'details': 'Project ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        test_reports = TestReport.objects.filter(
            project_id=pk).order_by('-created_at').first()

        if not test_reports:
            return Response({'details': 'No test report found for this project'}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        sheet = wb.active
        assert sheet is not None, "Failed"

        sheet.title = "Test Report"

        headers = ["Section", "Details"]

        sheet.append(headers)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style

        created_at_date = test_reports.created_at.replace(
            tzinfo=None).date() if test_reports.created_at else None
        created_at_str = created_at_date.strftime(
            '%Y-%m-%d') if created_at_date else None

        data = [
            ("Project", project_name),
            ("Total Test Cases", test_reports.total_test_cases),
            ("Passed Test Cases", test_reports.passed_test_cases),
            ("Failed Test Cases", test_reports.failed_test_cases),
            ("Defect Summary", str(test_reports.defect_summary)),
            ("Observations", test_reports.observations),
            ("Recommendations", test_reports.recommendations),
            ("Created At", created_at_str)
        ]

        for row in data:
            sheet.append(row)

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

            for cell in col:
                cell.alignment = Alignment(wrap_text=True)
                if cell.value:
                    cell.border = border_style

        current_time = datetime.now().strftime("%Y%m%d-%H%M%S")
        media_root = settings.MEDIA_ROOT
        download_path = os.path.join(media_root, 'download')
        os.makedirs(download_path, exist_ok=True)

        file_name = f'{project_name}_TestReport-{current_time}.xlsx'
        file_path = os.path.join(download_path, file_name)

        # Save the workbook to the specified path
        wb.save(file_path)

        # Return the response with the file path
        return Response({
            'message': 'File has been successfully saved.',
            'file_path': file_path,
        })
